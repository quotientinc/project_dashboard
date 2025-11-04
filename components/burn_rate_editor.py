"""
Burn Rate Editor Component
Displays Hours and Hours By Month sheets with what-if editing capability
"""

import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from utils.logger import get_logger

logger = get_logger(__name__)
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def show_burn_rate_editor(project, db_manager, processor):
    """Display the burn rate editor with Hours and Hours By Month sheets"""

    # Load data
    allocations_df = db_manager.get_allocations(project_id=project['id'])
    time_entries_by_month = db_manager.get_time_entries_by_month(project['id'])
    employees_df = db_manager.get_employees()  # Load all employees for name/role lookup

    if allocations_df.empty:
        st.warning("No allocations found for this project. Please add team members to the project first.")
        return

    # Initialize session state for what-if edits
    if 'burn_rate_edits' not in st.session_state:
        st.session_state.burn_rate_edits = {}

    if 'burn_rate_show_all_months' not in st.session_state:
        st.session_state.burn_rate_show_all_months = True

    # Initialize additional planning months (for extending timeline beyond project end date)
    if 'burn_rate_additional_months' not in st.session_state:
        st.session_state.burn_rate_additional_months = 0

    # Initialize reset counter for forcing data_editor to reset
    if 'burn_rate_reset_counter' not in st.session_state:
        st.session_state.burn_rate_reset_counter = 0

    if 'burn_rate_hours_df' not in st.session_state or st.session_state.get('burn_rate_project_id') != project['id']:
        # Build initial Hours sheet data
        st.session_state.burn_rate_hours_df = processor.build_hours_sheet_data(
            project,
            allocations_df,
            time_entries_by_month
        )
        st.session_state.burn_rate_project_id = project['id']
        st.session_state.burn_rate_edits = {}
        st.session_state.burn_rate_additional_months = 0
        st.session_state.burn_rate_reset_counter = 0

    # 1. Display Budget Summary with Export button
    display_budget_summary(st.session_state.burn_rate_hours_df, project, processor, employees_df)

    # 2. Display Hours By Month - Summary Table
    st.markdown("---")
    st.markdown("### 📊 Hours By Month - Employee Summary")
    st.markdown("""
    *Employee-level totals and variance calculated from the Hours sheet below.*

    **Formulas:**
    - **Target Hours** = FTE Target × 160 hours/month × Number of Months
    - **Over/Under** = Total Hours - Target Hours
    - **Total Cost** = Total Hours × Rate
    """)
    display_hours_by_month_summary(
        st.session_state.burn_rate_hours_df,
        project,
        processor,
        employees_df,
        db_manager
    )

    # 3. Display Hours By Month - Monthly Data Table
    st.markdown("---")
    st.markdown("### 📅 Hours By Month - Monthly Breakdown")

    st.markdown("""
    *Monthly hours and costs by employee.*

    **Formulas:**
    - **Hours** = Actual + Projected hours for the month
    - **Cost** = Hours × Rate
    """)
    display_hours_by_month_monthly(
        st.session_state.burn_rate_hours_df,
        project,
        st.session_state.burn_rate_show_all_months,
        processor,
        employees_df
    )

    # Show All Months checkbox at bottom right
    col1, col2 = st.columns([10, 1])
    with col2:
        st.session_state.burn_rate_show_all_months = st.toggle(
            "Show all months",
            value=st.session_state.burn_rate_show_all_months,
            key="show_all_months_hbm"
        )
        if st.session_state.burn_rate_show_all_months != st.session_state.get('prev_show_all_months', False):
            st.session_state.prev_show_all_months = st.session_state.burn_rate_show_all_months
            st.rerun()

    # 4. Display Hours sheet (editable) - MOVED TO BOTTOM
    st.markdown("---")
    st.markdown("### ✏️ Hours Sheet (Edit FTE and Actual Hours)")

    help_text = "*Click on FTE or Actual Hours cells to edit. Changes will NOT be saved to the database.*"
    if st.session_state.burn_rate_additional_months > 0:
        help_text += f" **({st.session_state.burn_rate_additional_months} planning month(s) added)**"
    st.markdown(help_text)

    st.markdown("""
    **Formulas:**
    - **Possible** = Working Days × 8 hours/day × FTE
    - **Projected** = Remaining Days × 8 hours/day × FTE
    - **Total** = Actual + Projected
    - **Total Hours** = Sum of all monthly Total values
    """)

    display_hours_sheet(
        st.session_state.burn_rate_hours_df,
        project,
        st.session_state.burn_rate_show_all_months,
        processor,
        allocations_df,
        db_manager,
        time_entries_by_month
    )

    # Show what-if indicator at bottom
    if st.session_state.burn_rate_edits:
        st.info(f"🔬 What-If Mode Active - {len(st.session_state.burn_rate_edits)} changes made (not saved to database)")


def display_budget_summary(hours_df, project, processor, employees_df):
    """Display budget summary metrics with export button"""

    if hours_df.empty:
        return

    # Build Hours By Month data to get summary
    hbm_df, summary = processor.build_hours_by_month_data(hours_df, project, employees_df)

    st.markdown("### 💰 Budget Summary")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric("Actual Cost", f"${summary.get('actual_cost', 0):,.2f}")

    with col2:
        st.metric("Current Funding", f"${summary.get('current_funding', 0):,.2f}")

    with col3:
        balance = summary.get('balance', 0)
        st.metric(
            "Balance",
            f"${balance:,.2f}",
            delta=f"${balance:,.2f}",
            delta_color="normal" if balance >= 0 else "inverse"
        )

    # Export button at bottom right
    col1, col2 = st.columns([10, 1])
    with col2:
        if st.button("📥 Export", width='stretch'):
            excel_file = export_to_excel(
                hours_df,
                project,
                processor,
                employees_df
            )
            st.download_button(
                label="Download",
                data=excel_file,
                file_name=f"burn_rate_{project['name'].replace(' ', '_')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch'
            )


def display_hours_by_month_summary(hours_df, project, processor, employees_df, db_manager):
    """Display Hours By Month summary table with editable Nominal FTE Target and Target Hours"""

    if hours_df.empty:
        st.info("No data available for Hours By Month sheet")
        return

    # Build Hours By Month data
    hbm_df, summary = processor.build_hours_by_month_data(hours_df, project, employees_df)

    if hbm_df.empty:
        st.info("No data available for Hours By Month calculation")
        return

    # Build summary DataFrame with fixed columns only
    summary_df = hbm_df[['employee_name', 'role', 'nominal_fte_target', 'target_hours', 'rate', 'over_under', 'total_hours', 'total_cost']].copy()

    # Recalculate Over/Under in case it needs to be updated (it's already calculated in build_hours_by_month_data with edited target hours)
    summary_df['over_under'] = summary_df['total_hours'] - summary_df['target_hours']

    # Store original values for change detection
    original_df = summary_df.copy()

    # Rename columns to be more user-friendly
    summary_df = summary_df.rename(columns={
        'employee_name': 'Employee',
        'role': 'Role',
        'nominal_fte_target': 'Nominal FTE Target',
        'target_hours': 'Target Hours',
        'rate': 'Rate',
        'over_under': 'Over/Under',
        'total_hours': 'Total Hours',
        'total_cost': 'Total Cost'
    })

    # Initialize session state for FTE edits and reset counter
    if 'burn_rate_fte_edits' not in st.session_state:
        st.session_state.burn_rate_fte_edits = {}

    if 'burn_rate_summary_reset_counter' not in st.session_state:
        st.session_state.burn_rate_summary_reset_counter = 0

    # Column configuration
    column_config = {
        'Employee': st.column_config.TextColumn(
            'Employee',
            disabled=True,
            width='medium'
        ),
        'Role': st.column_config.TextColumn(
            'Role',
            disabled=True,
            width='medium'
        ),
        'Nominal FTE Target': st.column_config.NumberColumn(
            'Nominal FTE Target',
            help='Employee FTE allocation (0.0-2.0)',
            disabled=True,
            format='%.2f',
            width='small'
        ),
        'Target Hours': st.column_config.NumberColumn(
            'Target Hours',
            help='Editable: Target hours for project duration',
            min_value=0.0,
            step=10.0,
            format='%.2f',
            width='small'
        ),
        'Rate': st.column_config.NumberColumn(
            'Rate',
            disabled=True,
            format='$%.2f',
            width='small'
        ),
        'Over/Under': st.column_config.NumberColumn(
            'Over/Under',
            disabled=True,
            format='%.2f',
            width='small'
        ),
        'Total Hours': st.column_config.NumberColumn(
            'Total Hours',
            disabled=True,
            format='%.2f',
            width='small'
        ),
        'Total Cost': st.column_config.NumberColumn(
            'Total Cost',
            disabled=True,
            format='$%.2f',
            width='medium'
        )
    }

    # Display as editable table with unique key that changes on reset
    edited_df = st.data_editor(
        summary_df,
        column_config=column_config,
        width='stretch',
        height=min(400, (len(summary_df) + 1) * 35 + 3),
        hide_index=True,
        key=f'hours_by_month_summary_editor_{st.session_state.burn_rate_summary_reset_counter}'
    )

    # Detect changes and store them in session state for recalculation
    changes = []
    need_rerun = False
    for idx in range(len(summary_df)):
        employee_name = edited_df.iloc[idx]['Employee']
        original_fte = summary_df.iloc[idx]['Nominal FTE Target']
        edited_fte = edited_df.iloc[idx]['Nominal FTE Target']
        original_target_hours = summary_df.iloc[idx]['Target Hours']
        edited_target_hours = edited_df.iloc[idx]['Target Hours']

        if original_fte != edited_fte or original_target_hours != edited_target_hours:
            # Check if this is a new change compared to session state
            existing_edit = st.session_state.burn_rate_fte_edits.get(employee_name, {})
            is_new_change = (
                existing_edit.get('fte') != edited_fte or
                existing_edit.get('target_hours') != edited_target_hours
            )

            # Store the edited values in session state
            st.session_state.burn_rate_fte_edits[employee_name] = {
                'fte': edited_fte,
                'target_hours': edited_target_hours
            }

            # Trigger rerun if this is a new change to update Over/Under
            if is_new_change:
                need_rerun = True

    # Trigger rerun BEFORE showing buttons if we detected new changes
    if need_rerun:
        st.rerun()

    # Calculate number of months for target hours calculation
    start_date = pd.to_datetime(project['start_date'])
    end_date = pd.to_datetime(project['end_date'])
    months = pd.date_range(
        start=start_date.replace(day=1),
        end=end_date + pd.DateOffset(months=1),
        freq='MS'
    )[:-1]
    num_months = len(months)

    # Build changes list from all edits in session state (for displaying buttons)
    # Compare the current table values with the original database values
    for idx in range(len(summary_df)):
        employee_name = edited_df.iloc[idx]['Employee']

        # Check if this employee has edits in session state
        if employee_name in st.session_state.burn_rate_fte_edits:
            edits = st.session_state.burn_rate_fte_edits[employee_name]

            # Get the original values (what would be shown without edits)
            original_fte = summary_df.iloc[idx]['Nominal FTE Target']
            original_target_hours_from_fte = original_fte * 160 * num_months

            # Get the edited values
            edited_fte = edits.get('fte', original_fte)
            edited_target_hours = edits.get('target_hours', original_target_hours_from_fte)

            # If there's a difference, add to changes
            if edited_fte != original_fte or edited_target_hours != original_target_hours_from_fte:
                changes.append({
                    'employee_name': employee_name,
                    'original_fte': original_fte,
                    'edited_fte': edited_fte,
                    'original_target_hours': original_target_hours_from_fte,
                    'edited_target_hours': edited_target_hours
                })

    # Show save/reset buttons if there are changes
    if changes:
        col1, col2, col3 = st.columns([6, 1, 1])

        with col1:
            st.info(f"ℹ️ {len(changes)} unsaved change(s) - Click Save to persist `Target Hours` changes to database or Reset to revert")

        with col2:
            if st.button("🔄 Reset", key="reset_fte_changes"):
                # Clear only the FTE edits session state (this only affects Hours By Month summary table)
                if 'burn_rate_fte_edits' in st.session_state:
                    del st.session_state.burn_rate_fte_edits
                # Increment the summary reset counter to force the data_editor to rebuild with original values
                st.session_state.burn_rate_summary_reset_counter += 1
                st.success("✅ Reset Hours By Month summary to database values!")
                st.rerun()

        with col3:
            if st.button("💾 Save", type="primary", key="save_fte_changes"):
                # Note: FTE values are now stored in allocations table (allocated_fte field)
                # Target Hours changes are stored separately in the burn_rate_fte_edits session state
                # This save button currently only updates target hours for display purposes
                # To persist FTE changes to allocations table, we would need to:
                # 1. Identify which allocations to update (by employee_id and project_id)
                # 2. Update allocated_fte in those allocation records

                st.warning("⚠️ FTE values are now stored per-allocation. To update FTE, use the Projects page to edit allocations.")
                st.success(f"✅ Target Hours changes applied for {len(changes)} employee(s) (display only)!")
                # Clear session state to reload fresh data
                if 'burn_rate_fte_edits' in st.session_state:
                    del st.session_state.burn_rate_fte_edits
                st.rerun()

    # Trigger rerun if we detected new changes to update the Over/Under column
    if need_rerun:
        st.rerun()


def display_hours_by_month_monthly(hours_df, project, show_all_months, processor, employees_df):
    """Display Hours By Month monthly data table (time-series columns only)"""

    if hours_df.empty:
        st.info("No data available for Hours By Month sheet")
        return

    # Build Hours By Month data
    hbm_df, summary = processor.build_hours_by_month_data(hours_df, project, employees_df)

    if hbm_df.empty:
        st.info("No data available for Hours By Month calculation")
        return

    # Get all months that exist in the hours_df (including planning months)
    # Find all month columns (those with format YYYY-MM in column names)
    month_cols = [col for col in hours_df.columns if '-' in col and col.split('_')[-1].count('-') == 1]

    # Extract unique month keys (YYYY-MM format)
    month_keys = set()
    for col in month_cols:
        parts = col.split('_')
        if len(parts) >= 2:
            month_key = parts[-1]  # Get YYYY-MM part
            month_keys.add(month_key)

    # Convert to datetime objects and sort
    months = sorted([pd.to_datetime(mk + '-01') for mk in month_keys])

    # Filter based on show_all_months if needed
    if not show_all_months:
        today = datetime.now()
        current_month = today.replace(day=1)
        months = [m for m in months if m >= pd.Timestamp(current_month)]

    # Build monthly DataFrame - start with employee name
    monthly_df = hbm_df[['employee_name']].copy()

    # Add month columns
    for month_date in months:
        month_key = month_date.strftime('%Y-%m')
        month_label = month_date.strftime('%b %Y')

        hours_col = f'hours_{month_key}'
        cost_col = f'cost_{month_key}'

        if hours_col in hbm_df.columns:
            monthly_df[f'{month_label} Hours'] = hbm_df[hours_col]
        if cost_col in hbm_df.columns:
            monthly_df[f'{month_label} Cost'] = hbm_df[cost_col]

    # Column configuration to pin employee column and format numbers
    column_config = {
        'employee_name': st.column_config.TextColumn(
            'Employee',
            pinned=True,
            width='medium'
        )
    }

    # Add column config for all hours and cost columns with proper formatting
    for col in monthly_df.columns:
        if 'Hours' in col:
            column_config[col] = st.column_config.NumberColumn(
                col,
                format='%.2f'
            )
        elif 'Cost' in col:
            column_config[col] = st.column_config.NumberColumn(
                col,
                format='$%.2f'
            )

    # Apply conditional formatting with alternating month shading
    def apply_formatting(val, col_name, month_index):
        if pd.isna(val):
            return ''

        styles = []

        # Alternating gray background for even months (0-indexed)
        if month_index % 2 == 0:
            styles.append('background-color: rgba(128, 128, 128, 0.2)')

        # Negative costs (red font) - override background for emphasis
        if 'Cost' in col_name and val < 0:
            styles.append('color: #FF0000')

        # Hours over thresholds (light red background) - override alternating background
        if 'Hours' in col_name:
            if val > 176:
                styles = ['background-color: #F4C7C3']  # Replace background
            elif val > 160:
                styles = ['background-color: #F4C7C3']  # Replace background

        return '; '.join(styles)

    # Format and style the DataFrame
    styled_df = monthly_df.style

    # Apply formatting to hours and cost columns
    month_index = 0
    for col in monthly_df.columns:
        if col == 'employee_name':
            continue

        if 'Hours' in col or 'Cost' in col:
            # Increment month_index only when we hit a new month (Hours column)
            if 'Hours' in col:
                current_month_index = month_index
                month_index += 1
            else:
                # Cost column uses the same index as its Hours column
                current_month_index = month_index - 1

            styled_df = styled_df.map(
                lambda val, cidx=current_month_index: apply_formatting(val, col, cidx),
                subset=[col]
            )

            # Format numbers
            if 'Hours' in col:
                styled_df = styled_df.format({col: '{:,.2f}'})
            elif 'Cost' in col:
                styled_df = styled_df.format({col: '${:,.2f}'})

    st.dataframe(
        styled_df,
        column_config=column_config,
        width='stretch',
        height=min(400, (len(monthly_df) + 1) * 35 + 3),
        hide_index=True
    )


def display_hours_sheet(hours_df, project, show_all_months, processor, allocations_df, db_manager, time_entries_by_month):
    """Display the Hours sheet with direct table editing using st.data_editor"""

    if hours_df.empty:
        st.info("No data available for Hours sheet")
        return

    # Get all months that exist in the dataframe (including planning months)
    # Find all month columns (those with format YYYY-MM in column names)
    month_cols = [col for col in hours_df.columns if '-' in col and col.split('_')[-1].count('-') == 1]

    # Extract unique month keys (YYYY-MM format)
    month_keys = set()
    for col in month_cols:
        parts = col.split('_')
        if len(parts) >= 2:
            month_key = parts[-1]  # Get YYYY-MM part
            month_keys.add(month_key)

    # Convert to datetime objects and sort
    months = sorted([pd.to_datetime(mk + '-01') for mk in month_keys])

    # Filter based on show_all_months if needed
    if not show_all_months:
        today = datetime.now()
        current_month = today.replace(day=1)
        months = [m for m in months if m >= pd.Timestamp(current_month)]

    # Build display DataFrame with formatted column names
    display_df = hours_df[['employee_name', 'role', 'rate', 'total_projected_hours']].copy()

    # Build column configuration for data_editor with pinned fixed columns
    column_config = {
        'employee_name': st.column_config.TextColumn('Employee', disabled=True, pinned=True, width='medium'),
        'role': st.column_config.TextColumn('Role', disabled=True, pinned=True, width='medium'),
        'rate': st.column_config.NumberColumn('Rate', format='$%.2f', disabled=True, pinned=True, width='small'),
        'total_projected_hours': st.column_config.NumberColumn('Total Hours', format='%.2f', disabled=True, pinned=True, width='small'),
    }

    # Add month columns with proper configuration
    for month_date in months:
        month_key = month_date.strftime('%Y-%m')
        month_label = month_date.strftime('%b %Y')

        # FTE column (editable)
        fte_col = f'fte_{month_key}'
        if fte_col in hours_df.columns:
            display_df[f'{month_label}_FTE'] = hours_df[fte_col]
            column_config[f'{month_label}_FTE'] = st.column_config.NumberColumn(
                f'{month_label} FTE',
                help='Editable: Full-Time Equivalent (0-2.0)',
                min_value=0.0,
                max_value=2.0,
                step=0.05,
                format='%.2f'
            )

        # Possible column (calculated, read-only)
        possible_col = f'possible_{month_key}'
        if possible_col in hours_df.columns:
            display_df[f'{month_label}_Possible'] = hours_df[possible_col]
            column_config[f'{month_label}_Possible'] = st.column_config.NumberColumn(
                f'{month_label} Possible',
                help='Calculated: Maximum possible hours',
                format='%.2f',
                disabled=True
            )

        # Actual column (editable)
        actual_col = f'actual_{month_key}'
        if actual_col in hours_df.columns:
            display_df[f'{month_label}_Actual'] = hours_df[actual_col]
            column_config[f'{month_label}_Actual'] = st.column_config.NumberColumn(
                f'{month_label} Actual',
                help='Editable: Actual hours worked',
                min_value=0.0,
                step=0.25,
                format='%.2f'
            )

        # Projected column (calculated, read-only)
        projected_col = f'projected_{month_key}'
        if projected_col in hours_df.columns:
            display_df[f'{month_label}_Projected'] = hours_df[projected_col]
            column_config[f'{month_label}_Projected'] = st.column_config.NumberColumn(
                f'{month_label} Projected',
                help='Calculated: Projected future hours',
                format='%.2f',
                disabled=True
            )

        # Total column (calculated, read-only)
        total_col = f'total_{month_key}'
        if total_col in hours_df.columns:
            display_df[f'{month_label}_Total'] = hours_df[total_col]
            column_config[f'{month_label}_Total'] = st.column_config.NumberColumn(
                f'{month_label} Total',
                help='Calculated: Actual + Projected',
                format='%.2f',
                disabled=True
            )

    # Apply styling to Total columns (read-only columns, so styling will work)
    # Using rgba with transparency to adapt to both light and dark themes
    def apply_total_style(val):
        return 'background-color: rgba(128, 128, 128, 0.2)'  # Semi-transparent gray that adapts to theme

    styled_df = display_df.style

    # Get Total column names (including total_projected_hours and monthly totals)
    total_cols = [col for col in display_df.columns if '_Total' in col or col == 'total_projected_hours']

    # Apply light gray background to Total columns
    if total_cols:
        try:
            # Try newer pandas map method
            styled_df = styled_df.map(apply_total_style, subset=total_cols)
        except AttributeError:
            # Fall back to older applymap method
            styled_df = styled_df.applymap(apply_total_style, subset=total_cols)

    # Display editable table
    edited_df = st.data_editor(
        styled_df,
        column_config=column_config,
        width='stretch',
        height=min(400, (len(display_df) + 1) * 35 + 3),
        hide_index=True,
        key='hours_sheet_editor'
    )

    # Detect and apply changes
    handle_hours_sheet_edits(display_df, edited_df, hours_df, months, processor)

    # Display editable working days and remaining days table
    st.markdown("**Configure Working Days & Remaining Days:**")
    st.caption("*Edit the values below to customize working days and remaining days for each month. Leave blank to use calculated defaults.*")
    display_days_editor(project, allocations_df, db_manager, processor, hours_df, months)

    # Add Month and Reset buttons at bottom right
    col1, col2, col3 = st.columns([10, 1, 1])
    with col2:
        if st.button("➕ Add Month", width='stretch', key="add_month_bottom"):
            # Add a planning month (extends beyond project end date)
            st.session_state.burn_rate_additional_months += 1
            add_planning_month(st.session_state.burn_rate_hours_df, project, processor)
            st.rerun()
    with col3:
        if st.button("🔄 Reset", width='stretch', key="reset_bottom"):
            st.session_state.burn_rate_edits = {}
            st.session_state.burn_rate_additional_months = 0
            st.session_state.burn_rate_hours_df = processor.build_hours_sheet_data(
                project,
                allocations_df,
                time_entries_by_month
            )
            # Clear the days config tracker so the days editor resets too
            if 'burn_rate_last_days_config' in st.session_state:
                st.session_state.burn_rate_last_days_config = {}
            # Increment reset counter to force data_editor to rebuild
            st.session_state.burn_rate_reset_counter += 1
            st.rerun()


def handle_hours_sheet_edits(original_df, edited_df, hours_df, months, processor):
    """Detect changes in the edited dataframe and recalculate dependent values"""

    changes_detected = False

    # Compare original and edited dataframes
    for idx in range(len(original_df)):
        employee_name = original_df.iloc[idx]['employee_name']
        employee_idx = hours_df[hours_df['employee_name'] == employee_name].index[0]

        for month_date in months:
            month_key = month_date.strftime('%Y-%m')
            month_label = month_date.strftime('%b %Y')

            # Check FTE changes
            fte_col = f'{month_label}_FTE'
            if fte_col in edited_df.columns:
                original_value = original_df.iloc[idx][fte_col]
                edited_value = edited_df.iloc[idx][fte_col]

                if original_value != edited_value:
                    # Update session state
                    edit_key = f"{employee_name}_{month_key}_FTE"
                    st.session_state.burn_rate_edits[edit_key] = edited_value

                    # Update hours_df
                    st.session_state.burn_rate_hours_df.at[employee_idx, f'fte_{month_key}'] = edited_value

                    # Recalculate dependent values
                    recalculate_month_values(employee_idx, month_key, processor)
                    changes_detected = True

            # Check Actual Hours changes
            actual_col = f'{month_label}_Actual'
            if actual_col in edited_df.columns:
                original_value = original_df.iloc[idx][actual_col]
                edited_value = edited_df.iloc[idx][actual_col]

                if original_value != edited_value:
                    # Update session state
                    edit_key = f"{employee_name}_{month_key}_Actual"
                    st.session_state.burn_rate_edits[edit_key] = edited_value

                    # Update hours_df
                    st.session_state.burn_rate_hours_df.at[employee_idx, f'actual_{month_key}'] = edited_value

                    # Recalculate dependent values
                    recalculate_month_values(employee_idx, month_key, processor)
                    changes_detected = True

    if changes_detected:
        st.rerun()


def recalculate_month_values(employee_idx, month_key, processor):
    """Recalculate dependent values for a specific employee and month"""

    year = int(month_key.split('-')[0])
    month_num = int(month_key.split('-')[1])
    days_info = processor.calculate_working_days(year, month_num)

    fte = st.session_state.burn_rate_hours_df.at[employee_idx, f'fte_{month_key}']
    actual = st.session_state.burn_rate_hours_df.at[employee_idx, f'actual_{month_key}']

    # Recalculate Possible
    possible = days_info['working_days'] * 8 * fte if fte <= 1 else fte
    st.session_state.burn_rate_hours_df.at[employee_idx, f'possible_{month_key}'] = possible

    # Recalculate Projected
    projected = days_info['remaining_days'] * 8 * fte if fte <= 1 else 0
    st.session_state.burn_rate_hours_df.at[employee_idx, f'projected_{month_key}'] = projected

    # Recalculate Total
    total = actual + projected
    st.session_state.burn_rate_hours_df.at[employee_idx, f'total_{month_key}'] = total

    # Recalculate total_projected_hours
    total_projected = 0
    for col in st.session_state.burn_rate_hours_df.columns:
        if col.startswith('total_') and '-' in col:
            total_projected += st.session_state.burn_rate_hours_df.at[employee_idx, col]
    st.session_state.burn_rate_hours_df.at[employee_idx, 'total_projected_hours'] = total_projected


def export_to_excel(hours_df, project, processor, employees_df):
    """Export Hours and Hours By Month sheets to Excel with formatting"""

    # Create a new workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create Hours sheet
    ws_hours = wb.create_sheet("Hours")

    # Create Hours By Month sheet
    ws_hbm = wb.create_sheet("Hours By Month")

    # Build Hours By Month data
    hbm_df, summary = processor.build_hours_by_month_data(hours_df, project, employees_df)

    # Write Hours sheet
    write_hours_sheet(ws_hours, hours_df, project)

    # Write Hours By Month sheet
    write_hours_by_month_sheet(ws_hbm, hbm_df, summary, project)

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file


def write_hours_sheet(ws, hours_df, project):
    """Write Hours sheet data to Excel worksheet"""

    # Parse project dates
    start_date = pd.to_datetime(project['start_date'])
    end_date = pd.to_datetime(project['end_date'])

    months = pd.date_range(
        start=start_date.replace(day=1),
        end=end_date + pd.DateOffset(months=1),
        freq='MS'
    )[:-1]

    # Write headers
    row = 1
    col = 1

    # Fixed columns
    ws.cell(row, col, "Name").font = Font(bold=True)
    ws.cell(row, col + 1, "Role").font = Font(bold=True)
    ws.cell(row, col + 2, "Rate").font = Font(bold=True)
    ws.cell(row, col + 3, "Total Projected Hours").font = Font(bold=True)

    col = 5

    # Month columns
    for month_date in months:
        month_label = month_date.strftime('%b %Y')
        ws.cell(row, col, month_label).font = Font(bold=True)
        ws.cell(row + 1, col, "FTE").font = Font(bold=True)
        ws.cell(row + 1, col + 1, "Possible").font = Font(bold=True)
        ws.cell(row + 1, col + 2, "Actual").font = Font(bold=True)
        ws.cell(row + 1, col + 3, "Projected").font = Font(bold=True)
        ws.cell(row + 1, col + 4, "Total").font = Font(bold=True)
        col += 5

    # Write data
    row = 3
    for _, emp in hours_df.iterrows():
        col = 1
        ws.cell(row, col, emp['employee_name'])
        ws.cell(row, col + 1, emp['role'])
        ws.cell(row, col + 2, emp['rate'])
        ws.cell(row, col + 3, emp['total_projected_hours'])

        col = 5
        for month_date in months:
            month_key = month_date.strftime('%Y-%m')
            ws.cell(row, col, emp.get(f'fte_{month_key}', 0))
            ws.cell(row, col + 1, emp.get(f'possible_{month_key}', 0))
            ws.cell(row, col + 2, emp.get(f'actual_{month_key}', 0))
            ws.cell(row, col + 3, emp.get(f'projected_{month_key}', 0))
            ws.cell(row, col + 4, emp.get(f'total_{month_key}', 0))
            col += 5

        row += 1


def write_hours_by_month_sheet(ws, hbm_df, summary, project):
    """Write Hours By Month sheet data to Excel worksheet"""

    if hbm_df.empty:
        return

    # Parse project dates
    start_date = pd.to_datetime(project['start_date'])
    end_date = pd.to_datetime(project['end_date'])

    months = pd.date_range(
        start=start_date.replace(day=1),
        end=end_date + pd.DateOffset(months=1),
        freq='MS'
    )[:-1]

    # Write headers
    row = 1
    col = 1

    headers = ['Name', 'Role', 'FTE Target', 'Target Hours', 'Rate', 'Over/Under', 'Total Hours', 'Total Cost']
    for i, header in enumerate(headers):
        ws.cell(row, col + i, header).font = Font(bold=True)

    col = len(headers) + 1

    # Month columns
    for month_date in months:
        month_label = month_date.strftime('%b %Y')
        ws.cell(row, col, month_label).font = Font(bold=True)
        ws.cell(row + 1, col, "Hours").font = Font(bold=True)
        ws.cell(row + 1, col + 1, "Cost").font = Font(bold=True)
        col += 2

    # Write data
    row = 3
    for _, emp in hbm_df.iterrows():
        col = 1
        ws.cell(row, col, emp['employee_name'])
        ws.cell(row, col + 1, emp['role'])
        ws.cell(row, col + 2, emp['nominal_fte_target'])
        ws.cell(row, col + 3, emp['target_hours'])
        ws.cell(row, col + 4, emp['rate'])
        ws.cell(row, col + 5, emp['over_under'])
        ws.cell(row, col + 6, emp['total_hours'])
        ws.cell(row, col + 7, emp['total_cost'])

        # Apply conditional formatting for total_hours
        if emp['total_hours'] > 2000:
            ws.cell(row, col + 6).fill = PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid")

        col = 9
        for month_date in months:
            month_key = month_date.strftime('%Y-%m')
            hours_val = emp.get(f'hours_{month_key}', 0)
            cost_val = emp.get(f'cost_{month_key}', 0)

            ws.cell(row, col, hours_val)
            ws.cell(row, col + 1, cost_val)

            # Apply conditional formatting
            if hours_val > 176:
                ws.cell(row, col).fill = PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid")
            elif hours_val > 160:
                ws.cell(row, col).fill = PatternFill(start_color="F4C7C3", end_color="F4C7C3", fill_type="solid")

            if cost_val < 0:
                ws.cell(row, col + 1).font = Font(color="FF0000")

            col += 2

        row += 1

    # Add summary rows
    row += 2
    ws.cell(row, 1, "Actual Cost").font = Font(bold=True)
    ws.cell(row, 2, summary.get('actual_cost', 0))

    row += 1
    ws.cell(row, 1, "Current Funding").font = Font(bold=True)
    ws.cell(row, 2, summary.get('current_funding', 0))

    row += 1
    ws.cell(row, 1, "Balance").font = Font(bold=True)
    ws.cell(row, 2, summary.get('balance', 0))
    if summary.get('balance', 0) < 0:
        ws.cell(row, 2).font = Font(color="FF0000")


def display_working_days_editor(project, allocations_df, db_manager, processor):
    """Display an editor for customizing working days and remaining days per month"""

    # Parse project dates
    start_date = pd.to_datetime(project['start_date'])
    end_date = pd.to_datetime(project['end_date'])

    # Generate month range
    months = pd.date_range(
        start=start_date.replace(day=1),
        end=end_date + pd.DateOffset(months=1),
        freq='MS'
    )[:-1]

    # Get current working days and remaining days from allocations
    # Group by month to get unique values
    days_by_month = {}
    for month_date in months:
        month_key = month_date.strftime('%Y-%m-%d')
        month_allocs = allocations_df[allocations_df['allocation_date'] == month_key]

        if not month_allocs.empty:
            # Take the first value (they should all be the same for a given month)
            working_days = int(month_allocs['working_days'].iloc[0]) if 'working_days' in month_allocs.columns and pd.notna(month_allocs['working_days'].iloc[0]) else None
            remaining_days = int(month_allocs['remaining_days'].iloc[0]) if 'remaining_days' in month_allocs.columns and pd.notna(month_allocs['remaining_days'].iloc[0]) else None
        else:
            working_days = None
            remaining_days = None

        # Calculate defaults if not set
        year = month_date.year
        month_num = month_date.month
        days_info = processor.calculate_working_days(year, month_num)

        if working_days is None:
            working_days = days_info['working_days']
        if remaining_days is None:
            remaining_days = days_info['remaining_days']

        days_by_month[month_date] = {
            'working_days': working_days,
            'remaining_days': remaining_days,
            'default_working_days': days_info['working_days'],
            'default_remaining_days': days_info['remaining_days']
        }

    # Display info message
    st.info("""
    **Working Days** define the total number of working days in each month.
    **Remaining Days** define how many working days are left in the month.

    - **Possible Hours** = Working Days × 8 hours/day × FTE
    - **Projected Hours** = Remaining Days × 8 hours/day × FTE

    By default, these are calculated automatically based on weekdays (Mon-Fri) and the current date, but you can override them here for planning purposes.
    """)

    # Build DataFrame for editing
    editor_data = []
    for month_date in months:
        month_label = month_date.strftime('%b %Y')
        data = days_by_month[month_date]

        is_working_customized = data['working_days'] != data['default_working_days']
        is_remaining_customized = data['remaining_days'] != data['default_remaining_days']

        editor_data.append({
            'month': month_label,
            'month_date': month_date,
            'default_working_days': data['default_working_days'],
            'working_days': data['working_days'],
            'default_remaining_days': data['default_remaining_days'],
            'remaining_days': data['remaining_days'],
            'is_customized': is_working_customized or is_remaining_customized
        })

    editor_df = pd.DataFrame(editor_data)

    # Display the editor
    st.markdown("**Edit Working Days & Remaining Days:**")

    edited_df = st.data_editor(
        editor_df[['month', 'default_working_days', 'working_days', 'default_remaining_days', 'remaining_days', 'is_customized']],
        column_config={
            'month': st.column_config.TextColumn('Month', disabled=True, width='medium'),
            'default_working_days': st.column_config.NumberColumn(
                'Default WD',
                help='Default working days (calculated)',
                disabled=True,
                width='small'
            ),
            'working_days': st.column_config.NumberColumn(
                'Working Days',
                help='Editable: Total working days in month',
                min_value=0,
                max_value=31,
                step=1,
                width='small'
            ),
            'default_remaining_days': st.column_config.NumberColumn(
                'Default RD',
                help='Default remaining days (calculated)',
                disabled=True,
                width='small'
            ),
            'remaining_days': st.column_config.NumberColumn(
                'Remaining Days',
                help='Editable: Remaining working days in month',
                min_value=0,
                max_value=31,
                step=1,
                width='small'
            ),
            'is_customized': st.column_config.CheckboxColumn(
                'Custom?',
                help='Indicates if customized',
                disabled=True,
                width='small'
            )
        },
        hide_index=True,
        width='stretch',
        key='working_days_editor'
    )

    # Detect changes
    changes = []
    for idx in range(len(editor_df)):
        working_changed = editor_df.iloc[idx]['working_days'] != edited_df.iloc[idx]['working_days']
        remaining_changed = editor_df.iloc[idx]['remaining_days'] != edited_df.iloc[idx]['remaining_days']

        if working_changed or remaining_changed:
            month_date = editor_df.iloc[idx]['month_date']
            new_working_days = edited_df.iloc[idx]['working_days']
            new_remaining_days = edited_df.iloc[idx]['remaining_days']
            changes.append((month_date, new_working_days, new_remaining_days))

    # Show save button if there are changes
    if changes:
        col1, col2 = st.columns([5, 1])
        with col2:
            if st.button("💾 Save Changes", type="primary"):
                # Update working_days and remaining_days in database for all allocations in the changed months
                for month_date, new_working_days, new_remaining_days in changes:
                    month_key = month_date.strftime('%Y-%m-%d')
                    month_allocs = allocations_df[allocations_df['allocation_date'] == month_key]

                    for _, alloc in month_allocs.iterrows():
                        db_manager.update_allocation(
                            alloc['id'],
                            {
                                'working_days': int(new_working_days),
                                'remaining_days': int(new_remaining_days)
                            }
                        )

                # Clear cached hours data to force rebuild
                if 'burn_rate_hours_df' in st.session_state:
                    del st.session_state.burn_rate_hours_df

                st.success(f"Updated days for {len(changes)} month(s)!")
                st.rerun()

        with col1:
            st.warning(f"⚠️ You have unsaved changes for {len(changes)} month(s). Click 'Save Changes' to apply them.")
    else:
        # Show reset to defaults button
        if any(editor_df['is_customized']):
            if st.button("🔄 Reset All to Defaults"):
                # Reset all working_days and remaining_days to calculated defaults
                for month_date in months:
                    month_key = month_date.strftime('%Y-%m-%d')
                    month_allocs = allocations_df[allocations_df['allocation_date'] == month_key]

                    # Calculate defaults
                    year = month_date.year
                    month_num = month_date.month
                    days_info = processor.calculate_working_days(year, month_num)

                    for _, alloc in month_allocs.iterrows():
                        db_manager.update_allocation(
                            alloc['id'],
                            {
                                'working_days': days_info['working_days'],
                                'remaining_days': days_info['remaining_days']
                            }
                        )

                # Clear cached hours data to force rebuild
                if 'burn_rate_hours_df' in st.session_state:
                    del st.session_state.burn_rate_hours_df

                st.success("Reset all days to defaults!")
                st.rerun()


def add_planning_month(hours_df, project, processor):
    """Add a planning month to the hours dataframe (extends beyond project end date)

    Returns:
        str: The new month key (YYYY-MM format) if successful, None otherwise
    """

    # Get the last month in the current hours_df
    # Find all month columns (those with format YYYY-MM in column names)
    month_cols = [col for col in hours_df.columns if '-' in col and col.split('_')[-1].count('-') == 1]

    if not month_cols:
        return None

    # Extract month keys (YYYY-MM format)
    month_keys = []
    for col in month_cols:
        parts = col.split('_')
        if len(parts) >= 2:
            month_key = parts[-1]  # Get YYYY-MM part
            if month_key not in month_keys:
                month_keys.append(month_key)

    # Find the latest month
    month_keys.sort()
    last_month_key = month_keys[-1]
    last_month_date = pd.to_datetime(last_month_key + '-01')

    # Add one month
    new_month_date = last_month_date + pd.DateOffset(months=1)
    new_month_key = new_month_date.strftime('%Y-%m')

    # Get working days for the new month
    year = new_month_date.year
    month_num = new_month_date.month
    days_info = processor.calculate_working_days(year, month_num)
    working_days = days_info['working_days']
    remaining_days = days_info['remaining_days']

    # Add columns for each employee in the dataframe
    for idx in range(len(hours_df)):
        # Get the last month's FTE value as default for the new month
        last_fte_col = f'fte_{last_month_key}'
        if last_fte_col in hours_df.columns:
            fte = hours_df.at[idx, last_fte_col]
        else:
            fte = 0.0

        # Calculate values for the new month
        # For planning months, we assume they are all in the future, so:
        # - Actual = 0 (no hours worked yet)
        # - Projected = full month's hours based on FTE
        # - Possible = working_days * 8 * FTE

        possible = working_days * 8 * fte if fte <= 1 else fte
        actual = 0.0
        projected = remaining_days * 8 * fte if fte <= 1 else fte
        total = actual + projected

        # Add new columns to the dataframe
        hours_df.at[idx, f'fte_{new_month_key}'] = fte
        hours_df.at[idx, f'working_days_{new_month_key}'] = working_days
        hours_df.at[idx, f'remaining_days_{new_month_key}'] = remaining_days
        hours_df.at[idx, f'possible_{new_month_key}'] = possible
        hours_df.at[idx, f'actual_{new_month_key}'] = actual
        hours_df.at[idx, f'projected_{new_month_key}'] = projected
        hours_df.at[idx, f'total_{new_month_key}'] = total

        # Update total_projected_hours
        total_projected = 0
        for col in hours_df.columns:
            if col.startswith('total_') and '-' in col:
                total_projected += hours_df.at[idx, col]
        hours_df.at[idx, 'total_projected_hours'] = total_projected

    return new_month_key


def display_days_editor(project, allocations_df, db_manager, processor, hours_df, months):
    """Display an editable 2-row table for working days and remaining days configuration"""

    # Get current working days and remaining days from allocations  or use defaults
    ref_data = {'Metric': 'Working Days'}
    ref_data2 = {'Metric': 'Remaining Days'}

    # Store month_date mapping for later updates
    month_date_map = {}

    for month_date in months:
        month_key = month_date.strftime('%Y-%m')
        month_label = month_date.strftime('%b %Y')
        month_alloc_key = month_date.strftime('%Y-%m-%d')
        month_date_map[month_label] = month_date

        # Try to get values from allocations first (database)
        month_allocs = allocations_df[allocations_df['allocation_date'] == month_alloc_key]

        if not month_allocs.empty:
            # Take the first value (they should all be the same for a given month)
            working_days = int(month_allocs['working_days'].iloc[0]) if 'working_days' in month_allocs.columns and pd.notna(month_allocs['working_days'].iloc[0]) else None
            remaining_days = int(month_allocs['remaining_days'].iloc[0]) if 'remaining_days' in month_allocs.columns and pd.notna(month_allocs['remaining_days'].iloc[0]) else None
        else:
            # No database records - use calculated defaults
            working_days = None
            remaining_days = None

        # Calculate defaults if not set
        year = month_date.year
        month_num = month_date.month
        days_info = processor.calculate_working_days(year, month_num)

        if working_days is None:
            working_days = days_info['working_days']
        if remaining_days is None:
            remaining_days = days_info['remaining_days']

        ref_data[month_label] = working_days
        ref_data2[month_label] = remaining_days

    # Create dataframe
    ref_df = pd.DataFrame([ref_data, ref_data2])

    # Column configuration
    column_config = {
        'Metric': st.column_config.TextColumn(
            'Metric',
            disabled=True,
            pinned=True,
            width='medium'
        )
    }

    # Add month columns as editable
    for month_date in months:
        month_label = month_date.strftime('%b %Y')
        if month_label in ref_df.columns:
            column_config[month_label] = st.column_config.NumberColumn(
                month_label,
                help='Editable: Enter custom value or leave for calculated default',
                min_value=0,
                max_value=31,
                step=1,
                width='small'
            )

    # Display as editable table (use reset counter in key to force rebuild on reset)
    edited_df = st.data_editor(
        ref_df,
        column_config=column_config,
        width='stretch',
        height=100,  # Small height for just 2 rows
        hide_index=True,
        key=f'days_editor_{st.session_state.burn_rate_reset_counter}'
    )

    # Initialize session state to track the last applied days configuration
    if 'burn_rate_last_days_config' not in st.session_state:
        st.session_state.burn_rate_last_days_config = {}

    # Detect changes and automatically recalculate (but don't save to database)
    changes = []
    changes_detected = False
    need_rerun = False

    for month_label in month_date_map.keys():
        if month_label in ref_df.columns and month_label in edited_df.columns:
            original_working = ref_df[month_label].iloc[0]
            edited_working = edited_df[month_label].iloc[0]
            original_remaining = ref_df[month_label].iloc[1]
            edited_remaining = edited_df[month_label].iloc[1]

            month_date = month_date_map[month_label]
            month_key = month_date.strftime('%Y-%m')
            month_alloc_key = month_date.strftime('%Y-%m-%d')

            # Check if this month exists in the database
            month_allocs = allocations_df[allocations_df['allocation_date'] == month_alloc_key]
            month_exists_in_db = not month_allocs.empty

            last_config_key = f"{month_key}_days"
            last_config = st.session_state.burn_rate_last_days_config.get(
                last_config_key,
                (original_working, original_remaining)
            )

            # Check if values differ from widget's edited values
            has_widget_changes = (original_working != edited_working or
                                 original_remaining != edited_remaining)

            # Check if there are applied changes in session state that differ from database
            has_applied_changes = (last_config != (original_working, original_remaining))

            # Check if this is an added month (exists in hours_df but not in database)
            is_added_month = not month_exists_in_db

            # Check if this is a NEW change (different from last applied config)
            is_new_change = last_config != (edited_working, edited_remaining)

            # Add to changes list if there are widget changes OR applied changes OR added month
            # This ensures Save button appears even when widget state is lost after reruns
            if has_widget_changes or has_applied_changes or is_added_month:
                # Use the most recent values (from widget if changed, else from last_config)
                if has_widget_changes:
                    final_working = edited_working
                    final_remaining = edited_remaining
                elif is_added_month:
                    # For added months, use the current ref_df values (not last_config)
                    final_working = original_working
                    final_remaining = original_remaining
                else:
                    final_working, final_remaining = last_config

                changes.append((month_date, final_working, final_remaining))
                changes_detected = True

            # If this is a NEW edit from the widget, update hours_df and trigger rerun
            if is_new_change and has_widget_changes:
                # Store the new configuration
                st.session_state.burn_rate_last_days_config[last_config_key] = (edited_working, edited_remaining)

                # Update hours_df with new working/remaining days for this month
                # This will automatically recalculate Possible and Projected hours
                for idx in range(len(hours_df)):
                    # Get FTE and actual for this employee and month
                    fte = hours_df.at[idx, f'fte_{month_key}']
                    actual = hours_df.at[idx, f'actual_{month_key}']

                    # Recalculate Possible with new working days
                    possible = edited_working * 8 * fte if fte <= 1 else fte
                    st.session_state.burn_rate_hours_df.at[idx, f'possible_{month_key}'] = possible
                    st.session_state.burn_rate_hours_df.at[idx, f'working_days_{month_key}'] = edited_working

                    # Recalculate Projected with new remaining days
                    projected = edited_remaining * 8 * fte if fte <= 1 else 0
                    st.session_state.burn_rate_hours_df.at[idx, f'projected_{month_key}'] = projected
                    st.session_state.burn_rate_hours_df.at[idx, f'remaining_days_{month_key}'] = edited_remaining

                    # Recalculate Total
                    total = actual + projected
                    st.session_state.burn_rate_hours_df.at[idx, f'total_{month_key}'] = total

                    # Recalculate total_projected_hours
                    total_projected = 0
                    for col in st.session_state.burn_rate_hours_df.columns:
                        if col.startswith('total_') and '-' in col:
                            total_projected += st.session_state.burn_rate_hours_df.at[idx, col]
                    st.session_state.burn_rate_hours_df.at[idx, 'total_projected_hours'] = total_projected

                need_rerun = True

    # Show save button if there are changes
    if changes:
        col1, col2 = st.columns([5, 1])
        with col2:
            if st.button("💾 Save", type="primary", key="save_days_changes"):
                # Update database with the changes
                for month_date, new_working_days, new_remaining_days in changes:
                    month_alloc_key = month_date.strftime('%Y-%m-%d')
                    month_allocs = allocations_df[allocations_df['allocation_date'] == month_alloc_key]

                    for _, alloc in month_allocs.iterrows():
                        db_manager.update_allocation(
                            alloc['id'],
                            {
                                'working_days': int(new_working_days),
                                'remaining_days': int(new_remaining_days)
                            }
                        )

                st.success(f"✅ Saved {len(changes)} month(s) to database!")
                st.rerun()

        with col1:
            st.info(f"ℹ️ {len(changes)} unsaved change(s) - Hours sheet updated, click Save to persist to database")

    # Trigger rerun only if we applied new changes
    if need_rerun:
        st.rerun()


def display_days_reference_table(hours_df, months):
    """Display a reference table showing working days and remaining days for each month"""

    # Build a single-row dataframe with working days and remaining days for each month
    ref_data = {'Metric': 'Working Days'}
    ref_data2 = {'Metric': 'Remaining Days'}

    for month_date in months:
        month_key = month_date.strftime('%Y-%m')
        month_label = month_date.strftime('%b %Y')

        # Get working days and remaining days from the first employee's data
        # (they should be the same for all employees in a given month)
        working_days_col = f'working_days_{month_key}'
        remaining_days_col = f'remaining_days_{month_key}'

        if working_days_col in hours_df.columns and not hours_df.empty:
            working_days = hours_df[working_days_col].iloc[0]
            ref_data[month_label] = int(working_days) if pd.notna(working_days) else 0
        else:
            ref_data[month_label] = 0

        if remaining_days_col in hours_df.columns and not hours_df.empty:
            remaining_days = hours_df[remaining_days_col].iloc[0]
            ref_data2[month_label] = int(remaining_days) if pd.notna(remaining_days) else 0
        else:
            ref_data2[month_label] = 0

    # Create dataframe
    ref_df = pd.DataFrame([ref_data, ref_data2])

    # Column configuration
    column_config = {
        'Metric': st.column_config.TextColumn(
            'Metric',
            pinned=True,
            width='medium'
        )
    }

    # Add month columns
    for month_date in months:
        month_label = month_date.strftime('%b %Y')
        if month_label in ref_df.columns:
            column_config[month_label] = st.column_config.NumberColumn(
                month_label,
                width='small'
            )

    # Display as a compact table
    st.dataframe(
        ref_df,
        column_config=column_config,
        width='stretch',
        height=100,  # Small height for just 2 rows
        hide_index=True
    )